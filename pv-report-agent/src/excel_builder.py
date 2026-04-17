"""
원시자료 분석 엑셀 생성기 — 모듈로 사용 가능
build_excel(files_dir, drug_code, drug_name) → bytes
"""
import io
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# 기본값 (단독 실행 시). src/excel_builder.py 기준 → drug-revision/docs/files
_DEFAULT_BASE = Path(__file__).parent.parent.parent / "docs" / "files"
_DEFAULT_DRUG_CODE = "201506668"
_DEFAULT_DRUG_NAME = "프로테조밉주3.5mg"

# ── 색상 정의 ──────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # 진한 파랑
SUB_FILL    = PatternFill("solid", fgColor="2E75B6")   # 중간 파랑
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")   # 연한 파랑 (홀짝 행)
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")

WHITE_FONT  = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
BOLD_FONT   = Font(name="맑은 고딕", bold=True, size=10)
NORMAL_FONT = Font(name="맑은 고딕", size=10)
SMALL_FONT  = Font(name="맑은 고딕", size=9)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

THIN = Side(style="thin", color="BFBFBF")
MED  = Side(style="medium", color="595959")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)


# ── 코드 매핑 ──────────────────────────────────────────────
RPT_TY_MAP = {"1": "자발적보고", "2": "시험/연구", "3": "기타", "4": "모름"}
PTNT_SEX_MAP = {"1": "남", "2": "여"}
PTNT_AGRDE_MAP = {
    "0": "태아", "1": "신생아(출생~28일)", "2": "영아(28일~24개월)",
    "3": "소아(24개월~12세)", "4": "청소년(12~19세)",
    "5": "성인(19~65세)", "6": "노인(65세이상)"
}
ADR_RESULT_MAP = {
    "1": "회복됨", "2": "회복중", "3": "회복안됨",
    "4": "후유증", "5": "치명적", "0": "알려지지않음"
}
DRUG_GB_MAP = {"1": "의심", "2": "병용", "3": "상호작용", "4": "비투여"}
DRUG_ACTION_MAP = {
    "1": "투여중지", "2": "투여량감소", "3": "투여량증가",
    "4": "투여량유지", "0": "모름", "9": "해당없음"
}
EVALT_RESULT_MAP = {
    "1": "확실함", "2": "상당히확실함", "3": "가능함",
    "4": "가능성적음", "5": "평가곤란", "6": "평가불가"
}
QCK_MAP = {"Y": "신속보고", "N": "일반보고"}

WHOART_SOC_MAP = {
    "01": "피부 및 피하조직 장애", "02": "근골격계 및 결합조직 장애",
    "03": "구강-위장관계 장애", "04": "중추/말초신경계 장애",
    "05": "자율신경계 장애", "06": "시각 장애",
    "07": "청각 및 전정 장애", "08": "심장 장애",
    "09": "혈관계 장애", "10": "호흡기계, 흉곽 및 종격 장애",
    "11": "적혈구 계통 장애", "12": "혈소판/출혈/응고 장애",
    "13": "백혈구 및 세망내피계 장애", "14": "간담도계 장애",
    "15": "대사 및 영양 장애", "16": "내분비 장애",
    "17": "비뇨기계 장애", "18": "생식기(여성) 장애",
    "19": "생식기(남성) 장애", "20": "신생아 및 영아 장애",
    "21": "전신계 장애", "22": "신생물",
    "23": "감염 및 기생충 침입", "24": "손상",
    "25": "선천성 장애", "26": "신생물 양성",
    "27": "심리적 장애", "28": "임신, 출산 및 주산기 상태",
    "29": "면역계 장애", "30": "신체 검사 결과 이상",
    "31": "의료 및 외과적 시술", "32": "사회 환경", "99": "기타",
}

SE_COLS = ["SE_DEATH", "SE_LIFE_MENACE", "SE_HSPTLZ_EXTN",
           "SE_FNCT_DGRD", "SE_ANMLY", "SE_ETC_IMPRTNC_SITTN"]
SE_LABEL = {"SE_DEATH": "사망", "SE_LIFE_MENACE": "생명위협",
            "SE_HSPTLZ_EXTN": "입원/연장", "SE_FNCT_DGRD": "기능저하/장애",
            "SE_ANMLY": "선천성기형", "SE_ETC_IMPRTNC_SITTN": "기타중요"}


def read_txt(fname, base=None, **kwargs):
    path = (base or _DEFAULT_BASE) / fname
    for enc in ["euc-kr", "utf-8"]:
        try:
            df = pd.read_csv(path, sep="|", encoding=enc, dtype=str, **kwargs)
            return df
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"인코딩 실패: {fname}")


def style_header(cell, sub=False):
    cell.fill = SUB_FILL if sub else HEADER_FILL
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def style_data(cell, alt=False, align=CENTER):
    cell.fill = ALT_FILL if alt else PatternFill()
    cell.font = SMALL_FONT
    cell.alignment = align
    cell.border = THIN_BORDER


def set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row, headers, sub=False):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        style_header(c, sub=sub)


def write_data_rows(ws, start_row, data: list[list], alt_start=False):
    for ri, row_data in enumerate(data):
        alt = (ri % 2 == (1 if alt_start else 0))
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=start_row + ri, column=ci, value=val)
            align = RIGHT if isinstance(val, (int, float)) else LEFT
            style_data(c, alt=alt, align=align)


# ═══════════════════════════════════════════════════════════
# 데이터 로드 & 전처리
# ═══════════════════════════════════════════════════════════
def load_data(base=None):
    b = base or _DEFAULT_BASE
    demo = read_txt("DEMO.txt", base=b)
    drug = read_txt("DRUG.txt", base=b)
    event = read_txt("EVENT.txt", base=b)
    if (b / "ASSESSMENT.txt").exists():
        assessment = read_txt("ASSESSMENT.txt", base=b)
    else:
        assessment = pd.DataFrame(columns=["KAERS_NO", "DRUG_SEQ", "ADR_SEQ", "EVALT_RESULT_CODE"])
    drug1 = read_txt("DRUG1.txt", base=b) if (b / "DRUG1.txt").exists() else pd.DataFrame(
        columns=["KAERS_NO", "DRUG_SEQ", "INGR_SEQ", "INGR_CD"]
    )
    drug2 = read_txt("DRUG2.txt", base=b) if (b / "DRUG2.txt").exists() else pd.DataFrame(
        columns=["KAERS_NO", "DRUG_SEQ", "DOSAGE_QTY", "DOSAGE_QTY_UNIT",
                 "DOSAGE_INTRVL", "DOSAGE_INTRVL_UNIT", "DOSAGE_START_DT", "DOSAGE_END_DT",
                 "DOSAGE_TERM", "DOSAGE_TERM_UNIT", "BNDE_LOT_NO",
                 "DRUG_SHAPE_ID", "DRUG_SHAPE_TXT", "DOSAGE_ROUTE_ID", "DOSAGE_ROUTE_TXT"]
    )
    drug3 = read_txt("DRUG3.txt", base=b) if (b / "DRUG3.txt").exists() else pd.DataFrame(
        columns=["KAERS_NO", "DRUG_SEQ", "EFFICACY_MEDDRA_KOR_NM",
                 "EFFICACY_MEDDRA_ENG_NM", "DSAS_CD", "DSAS_CD_VER"]
    )

    # 무효 건 제거
    if "REPRT_CHANGE_CD" in demo.columns:
        removed = (demo["REPRT_CHANGE_CD"] == "1").sum()
        demo = demo[demo["REPRT_CHANGE_CD"] != "1"].copy()
        print(f"  REPRT_CHANGE_CD=1 제거: {removed}건")

    # 코드 → 텍스트
    demo["RPT_TY_NM"]    = demo["RPT_TY"].map(RPT_TY_MAP).fillna(demo["RPT_TY"])
    demo["SEX_NM"]       = demo["PTNT_SEX"].map(PTNT_SEX_MAP).fillna("불상")
    demo["AGRDE_NM"]     = demo["PTNT_AGRDE"].map(PTNT_AGRDE_MAP).fillna("불상")
    demo["QCK_NM"]       = demo["QCK_RPT_YN"].map(QCK_MAP).fillna("일반보고")
    drug["DRUG_GB_NM"]   = drug["DRUG_GB"].map(DRUG_GB_MAP).fillna(drug["DRUG_GB"])
    drug["ACTION_NM"]    = drug["DRUG_ACTION"].map(DRUG_ACTION_MAP).fillna(drug["DRUG_ACTION"])
    event["RESULT_NM"]   = event["ADR_RESULT_CODE"].map(ADR_RESULT_MAP).fillna(event["ADR_RESULT_CODE"])
    event["SOC_NM"]      = event["WHOART_ARRN"].apply(
        lambda x: WHOART_SOC_MAP.get(str(x).strip()[:2].zfill(2), f"미확인({x})") if pd.notna(x) else "불상"
    )
    assessment["EVALT_NM"] = assessment["EVALT_RESULT_CODE"].map(EVALT_RESULT_MAP).fillna(assessment["EVALT_RESULT_CODE"])

    # 중대성
    for col in SE_COLS:
        if col not in event.columns:
            event[col] = np.nan
    event["IS_SERIOUS"] = event[SE_COLS].eq("Y").any(axis=1)
    event["SERIOUS_TYPES"] = event.apply(
        lambda r: ", ".join(SE_LABEL[c] for c in SE_COLS if r.get(c) == "Y") or "비중대", axis=1
    )

    # 날짜 포맷
    for col in ["RPT_DL_DT", "FIRST_OCCR_DT", "RECENT_OCCR_DT"]:
        if col in demo.columns:
            demo[col + "_FMT"] = demo[col].apply(
                lambda x: f"{x[:4]}-{x[4:6]}-{x[6:8]}" if pd.notna(x) and len(str(x)) == 8 else x
            )

    return demo, drug, event, assessment, drug1, drug2, drug3


def build_merged(demo, drug, event, assessment, drug_code=None):
    """대상 의약품 기준 통합 데이터프레임"""
    dc = drug_code or _DEFAULT_DRUG_CODE
    target_kaers = drug[
        (drug["DRUG_CD"] == dc) & (drug["DRUG_GB"].isin(["1", "2", "3"]))
    ]["KAERS_NO"].unique()

    demo_t = demo[demo["KAERS_NO"].isin(target_kaers)].copy()

    # 대상 의약품 행만 추출
    drug_t = drug[drug["KAERS_NO"].isin(target_kaers)].copy()
    drug_target = drug_t[drug_t["DRUG_CD"] == dc][["KAERS_NO", "DRUG_SEQ", "DRUG_GB_NM", "ACTION_NM"]].copy()
    drug_target.columns = ["KAERS_NO", "TARGET_DRUG_SEQ", "TARGET_DRUG_GB", "TARGET_DRUG_ACTION"]

    event_t = event[event["KAERS_NO"].isin(target_kaers)].copy()

    # 인과성: DRUG_SEQ 우선순위 (의심약물 우선)
    drug_prio = drug_t[drug_t["DRUG_CD"] == dc][["KAERS_NO", "DRUG_SEQ", "DRUG_GB"]].copy()
    drug_prio["PRIO"] = drug_prio["DRUG_GB"].apply(lambda x: 0 if x == "1" else 1)
    drug_prio = drug_prio.sort_values("PRIO").drop_duplicates("KAERS_NO")[["KAERS_NO", "DRUG_SEQ"]]

    assess_merged = assessment.merge(drug_prio, on=["KAERS_NO", "DRUG_SEQ"], how="inner")
    assess_for_event = assess_merged[["KAERS_NO", "ADR_SEQ", "EVALT_NM"]].copy()

    merged = event_t.merge(demo_t[["KAERS_NO", "RPT_DL_DT_FMT", "RPT_TY_NM", "QCK_NM",
                                    "SEX_NM", "AGRDE_NM", "PTNT_OCCURSYMT_AGE",
                                    "FIRST_OCCR_DT_FMT", "REPRT_CHANGE_CD"]], on="KAERS_NO", how="left")
    merged = merged.merge(drug_target, on="KAERS_NO", how="left")
    merged = merged.merge(assess_for_event, on=["KAERS_NO", "ADR_SEQ"], how="left")
    merged["EVALT_NM"] = merged["EVALT_NM"].fillna("[인과성 없음]")

    return merged


# ═══════════════════════════════════════════════════════════
# 시트 작성 함수들
# ═══════════════════════════════════════════════════════════

def write_summary_sheet(wb, demo, drug, event, merged, drug_name=None):
    ws = wb.create_sheet("요약 통계")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    dn = drug_name or _DEFAULT_DRUG_NAME
    row = 1
    # 제목
    c = ws.cell(row=row, column=2, value=f"■ {dn} 원시자료 분석 요약")
    c.font = Font(name="맑은 고딕", bold=True, size=14, color="1F4E79")
    c.alignment = LEFT
    row += 2

    sections = [
        ("전체 보고 현황", [
            ("전체 보고 건수 (DEMO)", len(demo)),
            ("유효 건수 (REPRT_CHANGE_CD≠1)", len(demo[demo["REPRT_CHANGE_CD"] != "1"]) if "REPRT_CHANGE_CD" in demo.columns else len(demo)),
            ("대상 의약품 관련 사례 수", merged["KAERS_NO"].nunique()),
            ("대상 의약품 관련 이상사례 수", len(merged)),
        ]),
        ("보고 유형", [
            (f"  {k}", v)
            for k, v in demo["RPT_TY_NM"].value_counts().items()
        ]),
        ("신속/일반 보고", [
            (f"  {k}", v)
            for k, v in demo["QCK_NM"].value_counts().items()
        ]),
        ("성별 현황 (대상 사례)", [
            (f"  {k}", v)
            for k, v in merged.drop_duplicates("KAERS_NO")["SEX_NM"].value_counts().items()
        ]),
        ("연령대 현황 (대상 사례)", [
            (f"  {k}", v)
            for k, v in merged.drop_duplicates("KAERS_NO")["AGRDE_NM"].value_counts().sort_index().items()
        ]),
        ("중대성 현황", [
            ("  중대한 이상사례", int(merged["IS_SERIOUS"].sum())),
            ("  비중대 이상사례", int((~merged["IS_SERIOUS"]).sum())),
        ]),
        ("SOC별 이상사례 건수 (상위 10)", [
            (f"  {k}", v)
            for k, v in merged["SOC_NM"].value_counts().head(10).items()
        ]),
    ]

    for section_title, rows in sections:
        # 섹션 헤더
        c = ws.cell(row=row, column=2, value=section_title)
        c.fill = SUB_FILL
        c.font = WHITE_FONT
        c.alignment = LEFT
        c.border = THIN_BORDER
        ws.cell(row=row, column=3).fill = SUB_FILL
        ws.cell(row=row, column=3).border = THIN_BORDER
        row += 1

        for label, val in rows:
            c1 = ws.cell(row=row, column=2, value=label)
            c2 = ws.cell(row=row, column=3, value=val)
            c1.font = SMALL_FONT
            c1.alignment = LEFT
            c1.border = THIN_BORDER
            c2.font = SMALL_FONT
            c2.alignment = RIGHT
            c2.border = THIN_BORDER
            if isinstance(val, int):
                c2.number_format = "#,##0"
            row += 1
        row += 1

    ws.freeze_panes = "B2"


def write_raw_demo(wb, demo):
    ws = wb.create_sheet("DEMO(기본정보)")
    ws.sheet_view.showGridLines = False

    display_cols = [
        ("KAERS_NO", "KAERS번호", 18),
        ("RPT_DL_DT_FMT", "보고접수일", 14),
        ("FIRST_OCCR_DT_FMT", "이상사례 발현일", 16),
        ("RECENT_OCCR_DT_FMT", "최근 발현일", 14),
        ("RPT_TY_NM", "보고유형", 14),
        ("QCK_NM", "신속/일반", 12),
        ("REPRT_CHANGE_CD", "변경코드", 10),
        ("PTNT_SEX", "성별코드", 10),
        ("SEX_NM", "성별", 8),
        ("PTNT_AGRDE", "연령대코드", 10),
        ("AGRDE_NM", "연령대", 18),
        ("PTNT_OCCURSYMT_AGE", "발현 당시 나이", 14),
        ("PTNT_WEIGHT", "체중(kg)", 10),
        ("PTNT_HEIGHT", "신장(cm)", 10),
        ("SENDER_TY", "보고자유형", 10),
    ]

    existing = [(c, n, w) for c, n, w in display_cols if c in demo.columns]
    cols = [c for c, n, w in existing]
    names = [n for c, n, w in existing]
    widths = [w for c, n, w in existing]

    set_col_widths(ws, [3] + widths)
    ws.column_dimensions["A"].width = 3

    write_header_row(ws, 1, [""] + names)
    for col_idx in range(len(names)):
        ws.cell(1, col_idx + 2)

    data = demo[cols].values.tolist()
    write_data_rows(ws, 2, data)
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30


def write_raw_drug(wb, drug, drug1, drug2, drug3, drug_code=None):
    ws = wb.create_sheet("DRUG(의약품정보)")
    ws.sheet_view.showGridLines = False

    # drug + drug2 병합
    merged = drug.merge(drug2, on=["KAERS_NO", "DRUG_SEQ"], how="left")
    merged = merged.merge(
        drug1.groupby(["KAERS_NO", "DRUG_SEQ"])["INGR_CD"].apply(lambda x: "/".join(x.dropna())).reset_index(),
        on=["KAERS_NO", "DRUG_SEQ"], how="left"
    )
    merged = merged.merge(
        drug3[["KAERS_NO", "DRUG_SEQ", "EFFICACY_MEDDRA_KOR_NM", "DSAS_CD"]],
        on=["KAERS_NO", "DRUG_SEQ"], how="left"
    )

    display_cols = [
        ("KAERS_NO", "KAERS번호", 18),
        ("DRUG_SEQ", "의약품순번", 10),
        ("DRUG_CD", "의약품코드", 14),
        ("DRUG_GB", "의약품구분\n코드", 10),
        ("DRUG_GB_NM", "의약품구분", 12),
        ("DRUG_ACTION", "처치코드", 10),
        ("ACTION_NM", "처치내용", 14),
        ("DOSAGE_QTY", "투여량", 10),
        ("DOSAGE_QTY_UNIT", "단위", 10),
        ("DOSAGE_INTRVL", "투여간격", 10),
        ("DOSAGE_START_DT", "투여시작일", 14),
        ("DOSAGE_END_DT", "투여종료일", 14),
        ("DOSAGE_ROUTE_TXT", "투여경로", 12),
        ("DRUG_SHAPE_TXT", "제형", 18),
        ("INGR_CD", "성분코드", 14),
        ("EFFICACY_MEDDRA_KOR_NM", "적응증", 20),
        ("DSAS_CD", "질병코드(ICD)", 14),
    ]

    existing = [(c, n, w) for c, n, w in display_cols if c in merged.columns]
    cols = [c for c, n, w in existing]
    names = [n for c, n, w in existing]
    widths = [w for c, n, w in existing]

    set_col_widths(ws, [3] + widths)
    ws.column_dimensions["A"].width = 3

    dc = drug_code or _DEFAULT_DRUG_CODE

    write_header_row(ws, 1, [""] + names)
    ws.row_dimensions[1].height = 35

    for ri, (_, row_data) in enumerate(merged[cols].iterrows()):
        is_target = (merged.iloc[ri]["DRUG_CD"] == dc)
        alt = ri % 2 == 1
        for ci, val in enumerate(row_data, 2):
            c = ws.cell(row=ri + 2, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=9, bold=is_target)
            c.alignment = LEFT if isinstance(val, str) else RIGHT
            c.border = THIN_BORDER
            if is_target:
                c.fill = PatternFill("solid", fgColor="FFF2CC")  # 노란
            elif alt:
                c.fill = ALT_FILL

    ws.freeze_panes = "B2"


def write_raw_event(wb, event):
    ws = wb.create_sheet("EVENT(이상사례)")
    ws.sheet_view.showGridLines = False

    display_cols = [
        ("KAERS_NO", "KAERS번호", 18),
        ("ADR_SEQ", "이상사례\n순번", 10),
        ("ADR_MEDDRA_KOR_NM", "이상사례명(한글)", 24),
        ("ADR_MEDDRA_ENG_NM", "이상사례명(영문)", 30),
        ("WHOART_ARRN", "WHOART\n코드", 12),
        ("SOC_NM", "기관계대분류(SOC)", 24),
        ("ADR_START_DT", "발현일", 14),
        ("ADR_END_DT", "종료일", 14),
        ("ADR_RESULT_CODE", "결과\n코드", 10),
        ("RESULT_NM", "이상사례 결과", 14),
        ("IS_SERIOUS", "중대성\n여부", 10),
        ("SERIOUS_TYPES", "중대성 유형", 22),
        ("SE_DEATH", "사망", 8),
        ("SE_LIFE_MENACE", "생명위협", 10),
        ("SE_HSPTLZ_EXTN", "입원/연장", 10),
        ("SE_FNCT_DGRD", "기능저하", 10),
        ("SE_ANMLY", "선천기형", 10),
        ("SE_ETC_IMPRTNC_SITTN", "기타중요", 10),
        ("CLNIC_FACT_CONFIRM_YN", "임상사실\n확인", 10),
    ]

    existing = [(c, n, w) for c, n, w in display_cols if c in event.columns]
    cols = [c for c, n, w in existing]
    names = [n for c, n, w in existing]
    widths = [w for c, n, w in existing]

    set_col_widths(ws, [3] + widths)
    ws.column_dimensions["A"].width = 3

    write_header_row(ws, 1, [""] + names)
    ws.row_dimensions[1].height = 35

    for ri, (_, row_data) in enumerate(event[cols].iterrows()):
        is_serious = (event.iloc[ri]["IS_SERIOUS"] == True)
        alt = ri % 2 == 1
        for ci, val in enumerate(row_data, 2):
            if isinstance(val, (bool, np.bool_)):
                val = "Y" if val else "N"
            c = ws.cell(row=ri + 2, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=9, bold=is_serious, color="C00000" if is_serious else "000000")
            c.alignment = CENTER if ci in [4, 5, 12, 13, 14, 15, 16, 17, 18] else LEFT
            c.border = THIN_BORDER
            if is_serious:
                c.fill = PatternFill("solid", fgColor="FCE4D6")
            elif alt:
                c.fill = ALT_FILL

    ws.freeze_panes = "B2"


def write_raw_assessment(wb, assessment):
    ws = wb.create_sheet("ASSESSMENT(인과성)")
    ws.sheet_view.showGridLines = False

    cols   = ["KAERS_NO", "DRUG_SEQ", "ADR_SEQ", "EVALT_RESULT_CODE", "EVALT_NM"]
    names  = ["KAERS번호", "의약품순번", "이상사례순번", "인과성코드", "인과성 평가"]
    widths = [18, 12, 14, 12, 18]

    set_col_widths(ws, [3] + widths)
    ws.column_dimensions["A"].width = 3

    write_header_row(ws, 1, [""] + names)
    ws.row_dimensions[1].height = 25

    data = assessment[[c for c in cols if c in assessment.columns]].values.tolist()
    write_data_rows(ws, 2, data)
    ws.freeze_panes = "B2"


def write_line_listing(wb, merged):
    ws = wb.create_sheet("Line Listing")
    ws.sheet_view.showGridLines = False

    headers = [
        "No.", "KAERS번호", "보고접수일", "이상사례\n발현일",
        "성별", "연령대", "나이", "보고유형", "신속/일반",
        "기관계대분류(SOC)", "이상사례명(한글)", "이상사례명(영문)",
        "이상사례 결과", "중대성", "중대성 유형",
        "의약품구분", "처치내용", "인과성 평가"
    ]
    widths = [6, 18, 14, 14, 8, 18, 8, 14, 12, 26, 22, 30, 14, 10, 20, 12, 14, 16]
    set_col_widths(ws, [3] + widths)
    ws.column_dimensions["A"].width = 3

    write_header_row(ws, 1, [""] + headers)
    ws.row_dimensions[1].height = 40

    col_map = [
        None,  # No. (generated)
        "KAERS_NO", "RPT_DL_DT_FMT", "ADR_START_DT",
        "SEX_NM", "AGRDE_NM", "PTNT_OCCURSYMT_AGE", "RPT_TY_NM", "QCK_NM",
        "SOC_NM", "ADR_MEDDRA_KOR_NM", "ADR_MEDDRA_ENG_NM",
        "RESULT_NM", None, "SERIOUS_TYPES",
        "TARGET_DRUG_GB", "TARGET_DRUG_ACTION", "EVALT_NM"
    ]

    for ri, (_, row_data) in enumerate(merged.iterrows()):
        alt = ri % 2 == 1
        is_serious = row_data.get("IS_SERIOUS", False)
        row_num = ri + 2
        for ci, col in enumerate(col_map, 2):
            if ci == 2:
                val = ri + 1  # No.
            elif col == "IS_SERIOUS" or col is None and ci == 15:
                val = "중대" if is_serious else "비중대"
            elif col:
                val = row_data.get(col, "")
            else:
                val = ""

            c = ws.cell(row=row_num, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=9)
            c.alignment = LEFT if ci > 11 else CENTER
            c.border = THIN_BORDER
            if is_serious:
                c.fill = PatternFill("solid", fgColor="FCE4D6")
            elif alt:
                c.fill = ALT_FILL

    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 40


def write_analysis_tables(wb, merged, demo):
    ws = wb.create_sheet("분석 테이블")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3

    row = 1

    def section_title(title, r):
        c = ws.cell(row=r, column=2, value=title)
        c.font = Font(name="맑은 고딕", bold=True, size=12, color="1F4E79")
        c.alignment = LEFT
        for col in range(2, 8):
            ws.cell(r, col).fill = PatternFill("solid", fgColor="D6E4F0")
            ws.cell(r, col).border = THIN_BORDER
        return r + 2

    # ── 1. 보고유형 × 신속/일반 교차표 ─────────────────────
    row = section_title("1. 보고유형별 보고 건수", row)
    target_demo = demo[demo["KAERS_NO"].isin(merged["KAERS_NO"].unique())]
    rpt_cross = pd.crosstab(
        target_demo["RPT_TY_NM"].fillna("모름"),
        target_demo["QCK_NM"].fillna("일반보고"),
        margins=True, margins_name="합계"
    ).reset_index()

    col_names = list(rpt_cross.columns)
    widths_t1 = [20] + [14] * (len(col_names) - 1)
    for i, w in enumerate(widths_t1, 2):
        ws.column_dimensions[get_column_letter(i)].width = w

    write_header_row(ws, row, [""] + col_names)
    ws.row_dimensions[row].height = 25
    row += 1
    for ri, r_data in enumerate(rpt_cross.values.tolist()):
        is_total = (r_data[0] == "합계")
        for ci, val in enumerate(r_data, 2):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=10, bold=is_total)
            c.alignment = CENTER if ci > 2 else LEFT
            c.border = THIN_BORDER
            if is_total:
                c.fill = PatternFill("solid", fgColor="E2EFDA")
        row += 1
    row += 2

    # ── 2. 성별 현황 ─────────────────────────────────────
    row = section_title("2. 성별 현황", row)
    sex_df = merged.drop_duplicates("KAERS_NO")["SEX_NM"].value_counts().reset_index()
    sex_df.columns = ["성별", "건수"]
    sex_df["비율(%)"] = (sex_df["건수"] / sex_df["건수"].sum() * 100).round(1)
    total_row = pd.DataFrame([["합계", sex_df["건수"].sum(), 100.0]], columns=sex_df.columns)
    sex_df = pd.concat([sex_df, total_row], ignore_index=True)

    write_header_row(ws, row, ["", "성별", "건수", "비율(%)"])
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.row_dimensions[row].height = 25
    row += 1
    for ri, r_data in enumerate(sex_df.values.tolist()):
        is_total = (r_data[0] == "합계")
        for ci, val in enumerate(r_data, 2):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=10, bold=is_total)
            c.alignment = CENTER if ci > 2 else LEFT
            c.border = THIN_BORDER
            if is_total:
                c.fill = PatternFill("solid", fgColor="E2EFDA")
        row += 1
    row += 2

    # ── 3. 연령대 현황 ────────────────────────────────────
    row = section_title("3. 연령대 현황", row)
    age_df = merged.drop_duplicates("KAERS_NO")["AGRDE_NM"].value_counts().reset_index()
    age_df.columns = ["연령대", "건수"]
    age_df["비율(%)"] = (age_df["건수"] / age_df["건수"].sum() * 100).round(1)
    total_row = pd.DataFrame([["합계", age_df["건수"].sum(), 100.0]], columns=age_df.columns)
    age_df = pd.concat([age_df, total_row], ignore_index=True)

    write_header_row(ws, row, ["", "연령대", "건수", "비율(%)"])
    ws.row_dimensions[row].height = 25
    row += 1
    for ri, r_data in enumerate(age_df.values.tolist()):
        is_total = (r_data[0] == "합계")
        for ci, val in enumerate(r_data, 2):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=10, bold=is_total)
            c.alignment = CENTER if ci > 2 else LEFT
            c.border = THIN_BORDER
            if is_total:
                c.fill = PatternFill("solid", fgColor="E2EFDA")
        row += 1
    row += 2

    # ── 4. SOC별 이상사례 현황 ────────────────────────────
    row = section_title("4. 기관계대분류(SOC)별 이상사례 현황", row)
    soc_df = merged["SOC_NM"].value_counts().reset_index()
    soc_df.columns = ["기관계대분류(SOC)", "건수"]
    soc_df["비율(%)"] = (soc_df["건수"] / soc_df["건수"].sum() * 100).round(1)
    total_row = pd.DataFrame([["합계", soc_df["건수"].sum(), 100.0]], columns=soc_df.columns)
    soc_df = pd.concat([soc_df, total_row], ignore_index=True)

    ws.column_dimensions["B"].width = 30
    write_header_row(ws, row, ["", "기관계대분류(SOC)", "건수", "비율(%)"])
    ws.row_dimensions[row].height = 25
    row += 1
    for ri, r_data in enumerate(soc_df.values.tolist()):
        is_total = (r_data[0] == "합계")
        for ci, val in enumerate(r_data, 2):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=10, bold=is_total)
            c.alignment = CENTER if ci > 2 else LEFT
            c.border = THIN_BORDER
            if is_total:
                c.fill = PatternFill("solid", fgColor="E2EFDA")
        row += 1
    row += 2

    # ── 5. PT별 이상사례 현황 (상위 20) ─────────────────
    row = section_title("5. 이상사례명(PT)별 현황 (전체)", row)
    pt_df = merged.groupby(["SOC_NM", "ADR_MEDDRA_KOR_NM", "ADR_MEDDRA_ENG_NM"]).size().reset_index(name="건수")
    pt_df = pt_df.sort_values("건수", ascending=False)
    pt_df["비율(%)"] = (pt_df["건수"] / pt_df["건수"].sum() * 100).round(1)
    total_row = pd.DataFrame([["", "합계", "", pt_df["건수"].sum(), 100.0]], columns=pt_df.columns)
    pt_df = pd.concat([pt_df, total_row], ignore_index=True)

    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 28
    write_header_row(ws, row, ["", "SOC", "이상사례명(한글)", "이상사례명(영문)", "건수", "비율(%)"])
    ws.row_dimensions[row].height = 25
    row += 1
    for ri, r_data in enumerate(pt_df.values.tolist()):
        is_total = (r_data[1] == "합계")
        for ci, val in enumerate(r_data, 2):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=10, bold=is_total)
            c.alignment = CENTER if ci > 4 else LEFT
            c.border = THIN_BORDER
            if is_total:
                c.fill = PatternFill("solid", fgColor="E2EFDA")
            elif ri % 2 == 1:
                c.fill = ALT_FILL
        row += 1
    row += 2

    # ── 6. 중대성 현황 ────────────────────────────────────
    row = section_title("6. 중대성 현황", row)
    serious_df = merged[["IS_SERIOUS", "SERIOUS_TYPES"]].copy()
    serious_summary = pd.DataFrame({
        "중대성 여부": ["중대한 이상사례(YES)", "비중대 이상사례(NO)", "합계"],
        "건수": [
            int(merged["IS_SERIOUS"].sum()),
            int((~merged["IS_SERIOUS"]).sum()),
            len(merged)
        ]
    })
    serious_summary["비율(%)"] = (serious_summary["건수"] / len(merged) * 100).round(1)
    serious_summary.loc[2, "비율(%)"] = 100.0

    write_header_row(ws, row, ["", "중대성 여부", "건수", "비율(%)"])
    ws.row_dimensions[row].height = 25
    row += 1
    for ri, r_data in enumerate(serious_summary.values.tolist()):
        is_total = (r_data[0] == "합계")
        for ci, val in enumerate(r_data, 2):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=10, bold=is_total)
            c.alignment = CENTER if ci > 2 else LEFT
            c.border = THIN_BORDER
            if is_total:
                c.fill = PatternFill("solid", fgColor="E2EFDA")
        row += 1
    row += 2

    # ── 7. 인과성 평가 현황 ──────────────────────────────
    row = section_title("7. 인과성 평가 현황", row)
    evalt_df = merged["EVALT_NM"].value_counts().reset_index()
    evalt_df.columns = ["인과성 평가", "건수"]
    evalt_df["비율(%)"] = (evalt_df["건수"] / evalt_df["건수"].sum() * 100).round(1)
    total_row = pd.DataFrame([["합계", evalt_df["건수"].sum(), 100.0]], columns=evalt_df.columns)
    evalt_df = pd.concat([evalt_df, total_row], ignore_index=True)

    write_header_row(ws, row, ["", "인과성 평가", "건수", "비율(%)"])
    ws.row_dimensions[row].height = 25
    row += 1
    for ri, r_data in enumerate(evalt_df.values.tolist()):
        is_total = (r_data[0] == "합계")
        for ci, val in enumerate(r_data, 2):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = Font(name="맑은 고딕", size=10, bold=is_total)
            c.alignment = CENTER if ci > 2 else LEFT
            c.border = THIN_BORDER
            if is_total:
                c.fill = PatternFill("solid", fgColor="E2EFDA")
        row += 1

    ws.freeze_panes = "B4"


def write_codebook_ref(wb):
    """코드집 참조 시트"""
    import openpyxl as xl
    ws = wb.create_sheet("코드 참조")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3

    title = ws.cell(1, 2, "■ 의약품부작용보고원시자료 코드 참조표")
    title.font = Font(name="맑은 고딕", bold=True, size=13, color="1F4E79")
    title.alignment = LEFT

    row = 3
    code_tables = [
        ("보고유형 (RPT_TY)", RPT_TY_MAP, 16, 20),
        ("성별 (PTNT_SEX)", PTNT_SEX_MAP, 14, 18),
        ("연령대 (PTNT_AGRDE)", PTNT_AGRDE_MAP, 14, 22),
        ("이상사례 결과 (ADR_RESULT_CODE)", ADR_RESULT_MAP, 14, 16),
        ("의약품구분 (DRUG_GB)", DRUG_GB_MAP, 14, 16),
        ("처치 (DRUG_ACTION)", DRUG_ACTION_MAP, 14, 18),
        ("인과성 평가 (EVALT_RESULT_CODE)", EVALT_RESULT_MAP, 14, 18),
    ]

    for table_name, mapping, w1, w2 in code_tables:
        c = ws.cell(row, 2, table_name)
        c.fill = SUB_FILL
        c.font = WHITE_FONT
        c.alignment = LEFT
        c.border = THIN_BORDER
        ws.cell(row, 3).fill = SUB_FILL
        ws.cell(row, 3).border = THIN_BORDER
        ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, w1)
        ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width or 0, w2)
        row += 1
        for code, name in mapping.items():
            ws.cell(row, 2, code).border = THIN_BORDER
            ws.cell(row, 2).font = SMALL_FONT
            ws.cell(row, 2).alignment = CENTER
            ws.cell(row, 3, name).border = THIN_BORDER
            ws.cell(row, 3).font = SMALL_FONT
            ws.cell(row, 3).alignment = LEFT
            row += 1
        row += 1

    # WHOART SOC 테이블
    c = ws.cell(row, 2, "WHO-ART SOC 코드 (WHOART_ARRN 앞 2자리)")
    c.fill = SUB_FILL
    c.font = WHITE_FONT
    c.alignment = LEFT
    c.border = THIN_BORDER
    ws.cell(row, 3).fill = SUB_FILL
    ws.cell(row, 3).border = THIN_BORDER
    row += 1
    for code, name in sorted(WHOART_SOC_MAP.items()):
        ws.cell(row, 2, code).border = THIN_BORDER
        ws.cell(row, 2).font = SMALL_FONT
        ws.cell(row, 2).alignment = CENTER
        ws.cell(row, 3, name).border = THIN_BORDER
        ws.cell(row, 3).font = SMALL_FONT
        ws.cell(row, 3).alignment = LEFT
        row += 1

    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 28


# ═══════════════════════════════════════════════════════════
# Word 연동 시트 (공유 집계 기반)
# ═══════════════════════════════════════════════════════════

def write_word_sync_sheet(wb, stats: dict, drug_name: str = ""):
    """Word 보고서에 기록된 것과 동일한 수치를 Excel에 렌더링 (집계 일관성 검증용)."""
    ws = wb.create_sheet("Word 연동 수치", 1)  # 요약 통계 다음에 위치
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14

    row = 1
    c = ws.cell(row=row, column=2, value=f"■ {drug_name} Word 보고서 연동 수치 (공유 집계)")
    c.font = Font(name="맑은 고딕", bold=True, size=14, color="1F4E79")
    c.alignment = LEFT
    row += 1
    note = ws.cell(row=row, column=2,
                   value="※ 이 시트의 수치는 Word 보고서와 동일한 aggregator.compute_aggregates()에서 유도됨")
    note.font = Font(name="맑은 고딕", italic=True, size=9, color="595959")
    note.alignment = LEFT
    row += 2

    def section_header(label, r):
        c = ws.cell(row=r, column=2, value=label)
        c.fill = SUB_FILL
        c.font = WHITE_FONT
        c.alignment = LEFT
        c.border = THIN_BORDER
        for col in range(3, 5):
            ws.cell(r, col).fill = SUB_FILL
            ws.cell(r, col).border = THIN_BORDER
        return r + 1

    def kv(r, label, val):
        c1 = ws.cell(row=r, column=2, value=label)
        c2 = ws.cell(row=r, column=3, value=val)
        c1.font = SMALL_FONT; c1.alignment = LEFT; c1.border = THIN_BORDER
        c2.font = SMALL_FONT; c2.alignment = RIGHT; c2.border = THIN_BORDER
        if isinstance(val, int):
            c2.number_format = "#,##0"
        return r + 1

    # 기본 수치
    row = section_header("기본 수치", row)
    row = kv(row, "이상사례 총 건수", stats.get("n_events", 0))
    row = kv(row, "보고 인원(사례) 수", stats.get("n_cases", 0))
    row = kv(row, "남성", stats.get("male", 0))
    row = kv(row, "여성", stats.get("female", 0))
    row = kv(row, "중대한 이상사례", stats.get("n_serious", 0))
    row = kv(row, "비중대 이상사례", stats.get("n_non_serious", 0))
    row = kv(row, "신속보고", stats.get("n_quick", 0))
    row += 1

    # 연령대
    age_counts = stats.get("age_counts")
    if age_counts is not None and len(age_counts) > 0:
        row = section_header("연령대별 (사례 단위)", row)
        for k, v in age_counts.items():
            row = kv(row, f"  {k}", int(v))
        row += 1

    # 보고유형 × 신속/일반
    rpt_cross = stats.get("rpt_cross")
    if rpt_cross is not None and len(rpt_cross) > 0:
        row = section_header("보고유형 × 신속/일반", row)
        write_header_row(ws, row, ["", "보고유형"] + list(rpt_cross.columns) + ["합계"])
        ws.row_dimensions[row].height = 22
        row += 1
        for idx, r_data in rpt_cross.iterrows():
            total = int(r_data.sum())
            cols = [idx] + [int(v) for v in r_data.values] + [total]
            for ci, val in enumerate(cols, 2):
                c = ws.cell(row=row, column=ci, value=val)
                c.font = SMALL_FONT
                c.alignment = LEFT if ci == 2 else RIGHT
                c.border = THIN_BORDER
            row += 1
        row += 1

    # SOC 요약
    soc_summary = stats.get("soc_summary")
    if soc_summary is not None and len(soc_summary) > 0:
        row = section_header("SOC(기관계대분류)별 건수", row)
        write_header_row(ws, row, ["", "SOC", "건수", "비율"])
        ws.row_dimensions[row].height = 22
        row += 1
        total_events = stats.get("n_events", 0) or 1
        for _, r_data in soc_summary.iterrows():
            pct = f"{r_data['건수'] / total_events * 100:.1f}%"
            for ci, val in enumerate([r_data["SOC_NM"], int(r_data["건수"]), pct], 2):
                c = ws.cell(row=row, column=ci, value=val)
                c.font = SMALL_FONT
                c.alignment = LEFT if ci == 2 else RIGHT
                c.border = THIN_BORDER
            row += 1

    ws.freeze_panes = "B4"


# ═══════════════════════════════════════════════════════════
# 공개 API
# ═══════════════════════════════════════════════════════════

def build_excel(
    files_dir: Path | str,
    drug_code: str,
    drug_name: str,
    shared_stats: dict | None = None,
) -> bytes:
    """
    files_dir: DEMO/DRUG/EVENT/ASSESSMENT.txt가 있는 디렉토리
    drug_code: 의약품 품목기준코드
    drug_name: 제품명 (엑셀 제목에 사용)
    shared_stats: aggregator.compute_aggregates() 결과. 있으면 Word 연동 검증 시트 추가.
    반환: xlsx bytes
    """
    base = Path(files_dir)
    demo, drug, event, assessment, drug1, drug2, drug3 = load_data(base=base)
    merged = build_merged(demo, drug, event, assessment, drug_code=drug_code)

    wb = Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, demo, drug, event, merged, drug_name=drug_name)
    if shared_stats is not None:
        write_word_sync_sheet(wb, shared_stats, drug_name=drug_name)
    write_analysis_tables(wb, merged, demo)
    write_line_listing(wb, merged)
    write_raw_demo(wb, demo)
    write_raw_drug(wb, drug, drug1, drug2, drug3, drug_code=drug_code)
    write_raw_event(wb, event)
    write_raw_assessment(wb, assessment)
    write_codebook_ref(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    out = Path(__file__).parent.parent / "data" / "output" / f"{_DEFAULT_DRUG_NAME}_원시자료분석.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_bytes(build_excel(_DEFAULT_BASE, _DEFAULT_DRUG_CODE, _DEFAULT_DRUG_NAME))
    print(f"✅ 완료: {out}")


if __name__ == "__main__":
    main()
